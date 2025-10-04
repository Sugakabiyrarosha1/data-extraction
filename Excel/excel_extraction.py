"""
Excel Data Extraction Module

This module provides robust, production-ready functions for extracting data from Excel files
into Pandas DataFrames with support for multiple sheets, data type inference, and comprehensive error handling.

Features:
- Multiple sheet support (single sheet, multiple sheets, or all sheets)
- Data type specification and enforcement
- Sheet information extraction without loading data
- Comprehensive error handling with meaningful messages
- Data validation and quality checks
- Flexible parsing options (header row, skip rows, etc.)
- Support for different Excel engines (openpyxl, xlrd)

Author: Data Extraction Toolkit
"""

import pandas as pd
import openpyxl
import os
from typing import Union, List, Dict, Any, Optional
import logging

# Configure logging for better debugging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def load_excel(path: str, 
               sheet_name: Union[str, int, List[Union[str, int]]] = 0,
               header_row: int = 0,
               skip_rows: int = 0,
               data_types: Optional[Dict[str, Any]] = None,
               engine: str = 'openpyxl',
               **kwargs) -> Union[pd.DataFrame, Dict[str, pd.DataFrame]]:
    """
    Load Excel data with advanced features including multiple sheet support.
    
    Args:
        path (str): Path to the Excel file
        sheet_name: Sheet name(s) to load. Can be:
            - int: Sheet index (0-based)
            - str: Sheet name
            - list: Multiple sheets (returns dict of DataFrames)
            - None: All sheets (returns dict of DataFrames)
        header_row (int): Row number to use as column headers
        skip_rows (int): Number of rows to skip from the beginning
        data_types (dict): Dictionary mapping column names to data types
        engine (str): Excel engine to use ('openpyxl' or 'xlrd')
        **kwargs: Additional arguments passed to pd.read_excel()
    
    Returns:
        pd.DataFrame or Dict[str, pd.DataFrame]: Loaded DataFrame(s)
    """
    try:
        # Check if file exists
        if not os.path.exists(path):
            raise FileNotFoundError(f"Excel file not found: {path}")
        
        logger.info(f"Loading Excel file: {path}")
        
        # Get sheet names if loading all sheets
        if sheet_name is None:
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            logger.info(f"Found {len(sheet_names)} sheets: {sheet_names}")
        
        # Load single sheet
        if isinstance(sheet_name, (str, int)):
            df = pd.read_excel(
                path,
                sheet_name=sheet_name,
                header=header_row,
                skiprows=skip_rows,
                dtype=data_types,
                engine=engine,
                **kwargs
            )
            
            # Data validation
            _validate_dataframe(df, f"{path} (sheet: {sheet_name})")
            logger.info(f"Successfully loaded sheet '{sheet_name}' with {len(df)} rows")
            return df
        
        # Load multiple sheets
        else:
            sheets_to_load = sheet_name if isinstance(sheet_name, list) else None
            dfs = pd.read_excel(
                path,
                sheet_name=sheets_to_load,
                header=header_row,
                skiprows=skip_rows,
                dtype=data_types,
                engine=engine,
                **kwargs
            )
            
            # Validate each sheet
            for sheet_name_key, df in dfs.items():
                _validate_dataframe(df, f"{path} (sheet: {sheet_name_key})")
                logger.info(f"Loaded sheet '{sheet_name_key}' with {len(df)} rows")
            
            return dfs
            
    except Exception as e:
        logger.error(f"Error loading Excel file {path}: {str(e)}")
        raise


def get_excel_sheet_info(path: str) -> Dict[str, Any]:
    """
    Get information about Excel file sheets without loading data.
    
    Args:
        path (str): Path to the Excel file
    
    Returns:
        dict: Information about sheets in the Excel file
    """
    try:
        workbook = openpyxl.load_workbook(path, read_only=True)
        sheet_info = {}
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_info[sheet_name] = {
                'max_row': sheet.max_row,
                'max_column': sheet.max_column,
                'dimensions': f"{sheet.max_row}x{sheet.max_column}"
            }
        
        workbook.close()
        return sheet_info
        
    except Exception as e:
        logger.error(f"Error reading Excel file info {path}: {str(e)}")
        raise


def load_excel_with_validation(path: str, 
                              sheet_name: Union[str, int] = 0,
                              required_columns: Optional[List[str]] = None,
                              **kwargs) -> pd.DataFrame:
    """
    Load Excel file with column validation.
    
    Args:
        path (str): Path to the Excel file
        sheet_name: Sheet name or index to load
        required_columns (list): List of required column names
        **kwargs: Additional arguments passed to load_excel()
    
    Returns:
        pd.DataFrame: Loaded and validated DataFrame
        
    Raises:
        ValueError: If required columns are missing
    """
    df = load_excel(path, sheet_name=sheet_name, **kwargs)
    
    if required_columns:
        missing_columns = set(required_columns) - set(df.columns)
        if missing_columns:
            raise ValueError(f"Missing required columns: {missing_columns}")
    
    return df


def export_to_excel(data: Union[pd.DataFrame, Dict[str, pd.DataFrame]], 
                   path: str,
                   sheet_name: str = 'Sheet1',
                   **kwargs) -> None:
    """
    Export DataFrame(s) to Excel file.
    
    Args:
        data: DataFrame or dictionary of DataFrames to export
        path (str): Output Excel file path
        sheet_name (str): Sheet name for single DataFrame
        **kwargs: Additional arguments passed to pd.ExcelWriter()
    """
    try:
        with pd.ExcelWriter(path, engine='openpyxl', **kwargs) as writer:
            if isinstance(data, pd.DataFrame):
                data.to_excel(writer, sheet_name=sheet_name, index=False)
                logger.info(f"Exported DataFrame to {path} (sheet: {sheet_name})")
            else:
                for sheet_name_key, df in data.items():
                    df.to_excel(writer, sheet_name=sheet_name_key, index=False)
                    logger.info(f"Exported DataFrame to {path} (sheet: {sheet_name_key})")
        
        logger.info(f"Successfully exported data to {path}")
        
    except Exception as e:
        logger.error(f"Error exporting to Excel file {path}: {str(e)}")
        raise


def _validate_dataframe(df: pd.DataFrame, source: str) -> None:
    """
    Validate DataFrame for common data quality issues.
    
    Args:
        df (pd.DataFrame): DataFrame to validate
        source (str): Source file path for logging
    """
    # Check for empty DataFrame
    if df.empty:
        raise ValueError(f"DataFrame from {source} is empty")
    
    # Check for completely empty columns
    empty_cols = df.columns[df.isnull().all()].tolist()
    if empty_cols:
        logger.warning(f"Found empty columns in {source}: {empty_cols}")
    
    # Check for duplicate column names
    if len(df.columns) != len(set(df.columns)):
        logger.warning(f"Found duplicate column names in {source}")
    
    # Log basic statistics
    logger.info(f"DataFrame shape: {df.shape}")
    logger.info(f"Memory usage: {df.memory_usage(deep=True).sum() / 1024**2:.2f} MB")


# Example usage and testing
if __name__ == "__main__":
    # Example usage with sample data
    import numpy as np
    from datetime import datetime, timedelta
    
    # Create sample data for demonstration
    sample_data = {
        'Sales': pd.DataFrame({
            'Date': pd.date_range('2024-01-01', periods=100, freq='D'),
            'Product': np.random.choice(['Laptop', 'Phone', 'Tablet', 'Monitor'], 100),
            'Quantity': np.random.randint(1, 10, 100),
            'Price': np.random.uniform(100, 2000, 100).round(2),
            'Salesperson': np.random.choice(['Alice', 'Bob', 'Charlie', 'Diana'], 100)
        }),
        'Products': pd.DataFrame({
            'Product_ID': range(1, 5),
            'Product_Name': ['Laptop', 'Phone', 'Tablet', 'Monitor'],
            'Category': ['Electronics', 'Electronics', 'Electronics', 'Electronics'],
            'Cost': [800, 400, 300, 200],
            'In_Stock': [True, True, False, True]
        })
    }
    
    # Save sample data to Excel file
    excel_path = 'sample_data.xlsx'
    export_to_excel(sample_data, excel_path)
    print(f"Created sample Excel file: {excel_path}")
    
    try:
        # Test 1: Get sheet information
        print("\n=== Testing Sheet Information ===")
        sheet_info = get_excel_sheet_info(excel_path)
        for sheet_name, info in sheet_info.items():
            print(f"  {sheet_name}: {info['dimensions']}")
        
        # Test 2: Load all sheets
        print("\n=== Testing Load All Sheets ===")
        all_sheets = load_excel(excel_path, sheet_name=None)
        for sheet_name, df in all_sheets.items():
            print(f"Sheet: {sheet_name}, Shape: {df.shape}")
        
        # Test 3: Load specific sheet with data types
        print("\n=== Testing Load Specific Sheet with Data Types ===")
        data_types = {
            'Date': 'datetime64[ns]',
            'Product': 'category',
            'Quantity': 'int32',
            'Price': 'float64'
        }
        df_sales = load_excel(excel_path, sheet_name='Sales', data_types=data_types)
        print(f"Sales data types after conversion:")
        print(df_sales.dtypes)
        
        # Test 4: Load with column validation
        print("\n=== Testing Load with Column Validation ===")
        required_columns = ['Date', 'Product', 'Price']
        df_validated = load_excel_with_validation(
            excel_path, 
            sheet_name='Sales', 
            required_columns=required_columns
        )
        print(f"Validated DataFrame shape: {df_validated.shape}")
        
    except Exception as e:
        print(f"Error during testing: {e}")
    
    finally:
        # Clean up sample file
        if os.path.exists(excel_path):
            os.remove(excel_path)
            print(f"\nCleaned up sample file: {excel_path}")
